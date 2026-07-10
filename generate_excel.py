import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def merge_and_style(ws, start_row, start_col, end_row, end_col, value, font, fill, alignment, border):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if r == start_row and c == start_col:
                cell.value = value
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border

def main():
    json_path = os.path.join(os.path.dirname(__file__), 'meritos_data.json')
    if not os.path.exists(json_path):
        print(f"Error: No se encontró el archivo JSON en {json_path}")
        return

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Primeros y Segundos Puestos"

    # Habilitar líneas de cuadrícula visibles
    ws.views.sheetView[0].showGridLines = True

    # 1. Configurar Título Principal (Columnas A a C)
    ws.merge_cells("A1:C1")
    ws['A1'] = "REPORTE DE PRIMEROS Y SEGUNDOS PUESTOS"
    ws['A1'].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep navy blue
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 2. Configurar Subtítulo (Columnas A a C)
    ws.merge_cells("A2:C2")
    ws['A2'] = "PROCESO DE ADMISIÓN 2026-I — ESCUELA DE POSGRADO UNPRG"
    ws['A2'].font = Font(name="Segoe UI", size=9.5, bold=True, color="D97706") # Gold
    ws['A2'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Fila 3 en blanco
    ws.row_dimensions[3].height = 12

    # 3. Encabezados de Tabla (Solo 3 columnas)
    headers = [
        "PROGRAMA ACADÉMICO", 
        "MÉRITO", 
        "POSTULANTE / INGRESANTE"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark slate
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 26

    # Estilos de borde
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), # Light slate border
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Rellenos y fuentes
    fill_program = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Soft slate-50
    fill_1st = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Gold light
    fill_2nd = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Silver light
    
    font_program = Font(name="Segoe UI", size=9.5, bold=True, color="1E3A8A") # Navy blue for program name
    font_postulante = Font(name="Segoe UI", size=10, bold=True, color="1F2937") # Bold dark gray name
    
    font_merit_1st = Font(name="Segoe UI", size=10, bold=True, color="B45309") # Amber
    font_merit_2nd = Font(name="Segoe UI", size=10, bold=True, color="475569") # Slate

    # Agrupar de forma correcta por programa
    grouped = defaultdict(list)
    for item in data:
        prog_name = f"{item['grado']} EN {item['programa']}"
        grouped[prog_name].append(item)

    # Mantener el orden original de aparición de los programas
    program_order = []
    seen = set()
    for item in data:
        prog_name = f"{item['grado']} EN {item['programa']}"
        if prog_name not in seen:
            seen.add(prog_name)
            program_order.append(prog_name)

    # Construir lista ordenada
    programs = []
    for prog_name in program_order:
        items_in_prog = grouped[prog_name]
        # Ordenar por mérito (1° Puesto primero, luego 2° Puestos)
        items_in_prog.sort(key=lambda x: x['merito'])
        programs.append((prog_name, items_in_prog))

    # 4. Escribir filas agrupadas
    row_idx = 5
    for prog_name, items in programs:
        num_rows = len(items)
        
        # A. Escribir y combinar Columna A (Programa Académico)
        merge_and_style(
            ws, 
            start_row=row_idx, 
            start_col=1, 
            end_row=row_idx + num_rows - 1, 
            end_col=1, 
            value=prog_name, 
            font=font_program, 
            fill=fill_program, 
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=thin_border
        )

        # B. Escribir Columnas B y C para cada ingresante del programa
        for offset, item in enumerate(items):
            curr_row = row_idx + offset
            ws.row_dimensions[curr_row].height = 24
            
            # Merit details
            merit_text = "1° Puesto" if item['merito'] == 1 else "2° Puesto"
            merit_font = font_merit_1st if item['merito'] == 1 else font_merit_2nd
            row_fill = fill_1st if item['merito'] == 1 else fill_2nd

            # Célula B (Mérito)
            cell_b = ws.cell(row=curr_row, column=2, value=merit_text)
            cell_b.font = merit_font
            cell_b.fill = row_fill
            cell_b.border = thin_border
            cell_b.alignment = Alignment(horizontal="center", vertical="center")

            # Célula C (Postulante / Ingresante)
            cell_c = ws.cell(row=curr_row, column=3, value=item['postulante'])
            cell_c.font = font_postulante
            cell_c.fill = row_fill
            cell_c.border = thin_border
            cell_c.alignment = Alignment(horizontal="left", vertical="center")

        row_idx += num_rows

    # 5. Configurar ancho de columnas
    column_widths = {
        'A': 75,   # PROGRAMA ACADÉMICO
        'B': 18,   # MÉRITO
        'C': 48    # POSTULANTE / INGRESANTE
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 6. Guardar archivo Excel
    output_filename = "reporte_primeros_segundos_puestos.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), output_filename)
    wb.save(output_path)
    
    # Limpieza
    try:
        os.remove(json_path)
    except Exception:
        pass

    print(f"Excel creado y formateado de forma premium en: {os.path.realpath(output_path)}")

if __name__ == "__main__":
    main()
